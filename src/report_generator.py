from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


REPORT_DIR = Path(__file__).parent.parent / "output" / "reports"


def generate_excel_report(df, kpis, region_summary, product_summary):
    """Generate a professional Excel sales report."""

    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    report_path = REPORT_DIR / "Sales_Report.xlsx"

    # ---------------------------------------------------------
    # 1. Create Excel file
    # ---------------------------------------------------------

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:

        df.to_excel(
            writer,
            sheet_name="Sales Data",
            index=False,
        )

        kpi_df = pd.DataFrame(
            {
                "Metric": [
                    "Total Sales",
                    "Total Profit",
                    "Total Quantity",
                    "Total Orders",
                    "Profit Margin",
                    "Average Order Value",
                ],
                "Value": [
                    kpis["total_sales"],
                    kpis["total_profit"],
                    kpis["total_quantity"],
                    kpis["total_orders"],
                    kpis["profit_margin"],
                    kpis["average_order_value"],
                ],
            }
        )

        kpi_df.to_excel(
            writer,
            sheet_name="KPI Summary",
            index=False,
        )

        region_summary.to_excel(
            writer,
            sheet_name="By Region",
        )

        product_summary.to_excel(
            writer,
            sheet_name="By Product",
        )

    # ---------------------------------------------------------
    # 2. Load workbook
    # ---------------------------------------------------------

    workbook = load_workbook(report_path)

    # ---------------------------------------------------------
    # 3. Create Dashboard
    # ---------------------------------------------------------

    dashboard = workbook.create_sheet("Dashboard", 0)

    dashboard["A1"] = "SALES PERFORMANCE DASHBOARD"
    dashboard["A1"].font = Font(
        bold=True,
        size=18,
    )

    dashboard.merge_cells("A1:D1")

    # KPI labels
    dashboard["A3"] = "Total Sales"
    dashboard["B3"] = kpis["total_sales"]

    dashboard["A4"] = "Total Profit"
    dashboard["B4"] = kpis["total_profit"]

    dashboard["A5"] = "Total Orders"
    dashboard["B5"] = kpis["total_orders"]

    dashboard["A6"] = "Profit Margin"
    dashboard["B6"] = kpis["profit_margin"] / 100

    dashboard["A7"] = "Average Order Value"
    dashboard["B7"] = kpis["average_order_value"]

    # KPI formatting
    for cell in ["A3", "A4", "A5", "A6", "A7"]:
        dashboard[cell].font = Font(bold=True)

    dashboard["B3"].number_format = '$#,##0.00'
    dashboard["B4"].number_format = '$#,##0.00'
    dashboard["B5"].number_format = '#,##0'
    dashboard["B6"].number_format = '0.00%'
    dashboard["B7"].number_format = '$#,##0.00'

    # ---------------------------------------------------------
    # 4. Format worksheets
    # ---------------------------------------------------------

    for sheet in workbook.worksheets:

        # Header formatting
        for cell in sheet[1]:
            cell.font = Font(
                bold=True,
            )

            cell.alignment = Alignment(
                horizontal="center",
            )

        # Auto-fit columns
        for column_cells in sheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )

            sheet.column_dimensions[
                column_letter
            ].width = min(max_length + 2, 40)

    # ---------------------------------------------------------
    # 5. Currency formatting
    # ---------------------------------------------------------

    for sheet_name in ["KPI Summary", "By Region", "By Product"]:

        sheet = workbook[sheet_name]

        for row in sheet.iter_rows():

            for cell in row:

                if isinstance(cell.value, (int, float)):

                    if (
                        "Sales" in str(
                            sheet.cell(
                                row=cell.row,
                                column=1,
                            ).value
                        )
                        or "Profit" in str(
                            sheet.cell(
                                row=cell.row,
                                column=1,
                            ).value
                        )
                    ):
                        cell.number_format = '$#,##0.00'

    # ---------------------------------------------------------
    # 6. Create Region Chart
    # ---------------------------------------------------------

    region_sheet = workbook["By Region"]

    chart = BarChart()

    chart.title = "Sales by Region"
    chart.y_axis.title = "Sales"
    chart.x_axis.title = "Region"

    data = Reference(
        region_sheet,
        min_col=2,
        min_row=1,
        max_row=region_sheet.max_row,
    )

    categories = Reference(
        region_sheet,
        min_col=1,
        min_row=2,
        max_row=region_sheet.max_row,
    )

    chart.add_data(
        data,
        titles_from_data=True,
    )

    chart.set_categories(categories)

    chart.height = 7
    chart.width = 12

    dashboard.add_chart(
        chart,
        "D3",
    )

    # ---------------------------------------------------------
    # 7. Create Product Chart
    # ---------------------------------------------------------

    product_sheet = workbook["By Product"]

    product_chart = BarChart()

    product_chart.title = "Sales by Product"
    product_chart.y_axis.title = "Sales"
    product_chart.x_axis.title = "Product"

    product_data = Reference(
        product_sheet,
        min_col=2,
        min_row=1,
        max_row=product_sheet.max_row,
    )

    product_categories = Reference(
        product_sheet,
        min_col=1,
        min_row=2,
        max_row=product_sheet.max_row,
    )

    product_chart.add_data(
        product_data,
        titles_from_data=True,
    )

    product_chart.set_categories(
        product_categories
    )

    product_chart.height = 7
    product_chart.width = 12

    dashboard.add_chart(
        product_chart,
        "D18",
    )

    # ---------------------------------------------------------
    # 8. Freeze panes
    # ---------------------------------------------------------

    workbook["Sales Data"].freeze_panes = "A2"
    workbook["By Region"].freeze_panes = "A2"
    workbook["By Product"].freeze_panes = "A2"

    # ---------------------------------------------------------
    # 9. Save workbook
    # ---------------------------------------------------------

    workbook.save(report_path)

    print(
        f"Professional Excel report created: {report_path}"
    )

    return report_path